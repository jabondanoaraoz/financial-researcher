"""
DCF Model Sheet  (v3 - full rewrite)
Full IB-style DCF with Excel formulas for all projections and derived cells.
Yellow input cells are user-editable; WACC and all projections are formulaic.
Single combined table: historical cols C-G + projection cols H-L in the same rows.

"""

import math
from openpyxl.utils import get_column_letter as _gcl

from ..styles import (
    _f, fill, merge, wc, sec_hdr, col_hdr, spacer, hide_gridlines,
    NAVY_DARK, NAVY_MED, BLUE_ACC, GREEN, AMBER, RED,
    BLUE_TINT, GRAY_LIGHT, WHITE, BLACK, DARK_GRAY, MID_GRAY,
    SIGNAL_BG,
    AL_L, AL_LI, AL_C, AL_R, AL_W,
    BORDER_ALL,
)

# Column layout: B=label, C-G=5 historical years, H-L=5 projection years
_COLS = {
    "A": 2,   # spacer
    "B": 24,  # label
    "C": 12,  # FY hist 1 (oldest)
    "D": 12,
    "E": 12,
    "F": 12,
    "G": 12,  # FY hist 5 (most recent)
    "H": 12,  # Proj yr 1
    "I": 12,
    "J": 12,
    "K": 12,
    "L": 12,  # Proj yr 5
    "M": 2,   # spacer
}
C1        = 2                       # col B
N_HIST    = 5
N_PROJ    = 5
HIST_START = C1 + 1                 # col C = 3
PROJ_START = C1 + 1 + N_HIST       # col H = 8
CE         = PROJ_START + N_PROJ - 1  # col L = 12
INPUT_COL  = C1 + 1                 # col C = 3 (for inputs section)

INPUT_YELLOW = "FFFF99"


def _a(row, col, ar=False, ac=False):
    c = ("$" if ac else "") + _gcl(col)
    r = ("$" if ar else "") + str(row)
    return c + r


def _abs(row, col):
    return _a(row, col, ar=True, ac=True)


# ── Data helpers ──────────────────────────────────────────────────────────────

def _get(df, row_key, col_idx=0):
    if df is None or df.empty or row_key not in df.index:
        return None
    try:
        v = float(df.iloc[df.index.get_loc(row_key), col_idx])
        return None if math.isnan(v) else v
    except (TypeError, ValueError, IndexError):
        return None


def _reversed_series(df, row_key, n=5):
    """Return up to n values oldest-first (API gives newest-first)."""
    if df is None or df.empty or row_key not in df.index:
        return [None] * n
    vals = []
    for v in df.loc[row_key].values[:n]:
        try:
            fv = float(v)
            vals.append(None if math.isnan(fv) else fv)
        except (TypeError, ValueError):
            vals.append(None)
    while len(vals) < n:
        vals.append(None)
    vals.reverse()
    return vals


def _newest_series(df, row_key, n=5):
    """Return up to n values newest-first."""
    vals = _reversed_series(df, row_key, n)
    vals.reverse()
    return vals


def _year_labels(df, n=5):
    """Return FY year labels oldest-first."""
    if df is None or df.empty:
        return [f"FY{2020+i}" for i in range(n)]
    labels = []
    for col in list(df.columns)[:n]:
        try:
            labels.append(f"FY{col.year}")
        except AttributeError:
            labels.append(str(col)[:6])
    while len(labels) < n:
        labels.insert(0, "—")
    labels.reverse()
    return labels


def _safe_cagr(revenues_newest_first):
    clean = [v for v in revenues_newest_first if v and v > 0]
    if len(clean) < 2:
        return 8.5
    n = len(clean) - 1
    cagr = (clean[0] / clean[-1]) ** (1.0 / n) - 1
    return round(min(max(cagr * 100, 0.0), 40.0), 1)


def _safe_pct(numerator, denominator, fallback=0.0):
    if numerator is None or denominator is None or denominator == 0:
        return fallback
    return round((numerator / denominator) * 100, 1)


# ── Sheet builder ─────────────────────────────────────────────────────────────

def build(wb, result):
    ws = wb.create_sheet("DCF Model")
    hide_gridlines(ws)
    ws.sheet_properties.tabColor = GREEN

    for col, w in _COLS.items():
        ws.column_dimensions[col].width = w

    ticker    = result.get("ticker", "")
    km        = (result.get("company_data") or {}).get("key_metrics") or {}
    fins      = result.get("financials") or {}
    risk_free = result.get("risk_free_rate") or 0.043
    price     = km.get("current_price")

    inc = fins.get("income_statement")
    cf  = fins.get("cash_flow")
    bs  = fins.get("balance_sheet")

    # ── Pre-compute seed values for input cells ───────────────────────────────
    rev_series   = _newest_series(inc, "Total Revenue",    N_HIST)
    gp_series    = _newest_series(inc, "Gross Profit",     N_HIST)
    ebitda_series= _newest_series(inc, "EBITDA",           N_HIST)
    da_series    = _newest_series(cf,  "Depreciation And Amortization", N_HIST)
    capex_series = _newest_series(cf,  "Capital Expenditure", N_HIST)
    ie_series    = _newest_series(inc, "Interest Expense", N_HIST)
    ca_series    = _newest_series(bs,  "Current Assets",   N_HIST)
    cl_series    = _newest_series(bs,  "Current Liabilities", N_HIST)

    rev0     = rev_series[0]    # most recent (index 0 = newest)
    gp0      = gp_series[0]
    ebitda0  = ebitda_series[0]
    da0      = da_series[0]
    capex0   = capex_series[0]
    ie0      = ie_series[0]
    debt0    = _get(bs, "Total Debt", 0)
    equity0  = _get(bs, "Stockholders Equity", 0)
    cash0    = _get(bs, "Cash Cash Equivalents And Short Term Investments", 0)
    shares   = km.get("shares_outstanding")

    cagr_seed          = _safe_cagr(rev_series)
    gm_seed            = _safe_pct(gp0, rev0, 40.0)
    ebitda_margin_seed = _safe_pct(ebitda0, rev0, 20.0)
    da_pct_seed        = _safe_pct(da0, rev0, 3.0)
    capex_pct_seed     = _safe_pct(abs(capex0) if capex0 else None, rev0, 4.0)
    beta_seed          = km.get("beta") or 1.0
    rf_seed            = round(risk_free * 100, 2)

    ie_abs = abs(ie0) if ie0 else None
    kd_seed = round((ie_abs / debt0) * 100, 1) if (ie_abs and debt0 and debt0 > 0) else 4.0

    V = (equity0 or 0) + (debt0 or 0)
    dw_seed = round((debt0 / V) * 100, 1) if (debt0 and V and V > 0) else 30.0

    rows = {}
    inp  = {}

    r = 1
    spacer(ws, r, 8); r += 1

    # Title
    ws.row_dimensions[r].height = 28
    wc(ws, r, C1, f"  DCF MODEL  -  {ticker}",
       font=_f(14, True, WHITE), bg=NAVY_DARK, align=AL_L)
    merge(ws, r, C1, r, CE)
    r += 1
    spacer(ws, r, 4); r += 1

    # ── SECTION 1: Model Inputs ───────────────────────────────────────────────
    sec_hdr(ws, r, "Model Assumptions  (yellow cells are editable)", C1, CE); r += 1

    def _inp_row(label, value, note="", is_wacc=False):
        nonlocal r
        ws.row_dimensions[r].height = 16
        bg_lbl = NAVY_DARK if is_wacc else BLUE_TINT
        bg_val = NAVY_DARK if is_wacc else INPUT_YELLOW
        ft_lbl = _f(9, True, WHITE) if is_wacc else _f(9, True)
        ft_val = _f(9, True, WHITE) if is_wacc else _f(9, True)
        wc(ws, r, C1, label, font=ft_lbl, bg=bg_lbl, align=AL_LI, border=BORDER_ALL)
        cell = ws.cell(row=r, column=INPUT_COL, value=value)
        cell.font = ft_val; cell.fill = fill(bg_val)
        cell.alignment = AL_R; cell.border = BORDER_ALL; cell.number_format = "0.00"
        if note:
            wc(ws, r, INPUT_COL + 1, note,
               font=_f(8, False, DARK_GRAY if not is_wacc else WHITE),
               bg=GRAY_LIGHT if not is_wacc else NAVY_DARK, align=AL_L, border=BORDER_ALL)
            merge(ws, r, INPUT_COL + 1, r, CE)
        this_row = r
        r += 1
        return this_row

    inp["cagr"]          = _inp_row("Revenue CAGR (%)",           cagr_seed,            "Historical revenue CAGR (adjust for sensitivity analysis)")
    inp["gross_margin"]  = _inp_row("Gross Margin (%)",           gm_seed,              "Gross Profit / Revenue (most recent year)")
    inp["ebitda_margin"] = _inp_row("EBITDA Margin (%)",          ebitda_margin_seed,   "EBITDA / Revenue (most recent year)")
    inp["da_pct"]        = _inp_row("D&A (% Revenue)",            da_pct_seed,          "Depreciation & Amortization / Revenue")
    inp["capex_pct"]     = _inp_row("CapEx (% Revenue)",          capex_pct_seed,       "CapEx / Revenue (absolute value)")
    inp["nwc_pct"]       = _inp_row("ΔNWC (% ΔRevenue)",          1.5,                  "Change in Net Working Capital as % of Revenue change")
    inp["tax_rate"]      = _inp_row("Tax Rate (%)",               21.0,                 "US corporate tax rate")
    inp["terminal_g"]    = _inp_row("Terminal Growth Rate (%)",   3.0,                  "Long-term nominal GDP growth")
    inp["beta"]          = _inp_row("Beta",                       round(beta_seed, 3),  "Beta vs S&P 500")
    inp["rf_rate"]       = _inp_row("Risk-Free Rate (%)",         rf_seed,              "Treasury 10Y (FRED)")
    inp["erp"]           = _inp_row("Equity Risk Premium (%)",    5.5,                  "Equity Risk Premium (Damodaran)")
    inp["debt_cost"]     = _inp_row("Cost of Debt (%)",           kd_seed,              "Interest Expense / Total Debt")
    inp["debt_weight"]   = _inp_row("Debt Weight D/V (%)",        dw_seed,              "Debt / (Debt + Equity)")

    # WACC formula row
    dw_abs  = _abs(inp["debt_weight"], INPUT_COL)
    rf_abs  = _abs(inp["rf_rate"],     INPUT_COL)
    b_abs   = _abs(inp["beta"],        INPUT_COL)
    erp_abs = _abs(inp["erp"],         INPUT_COL)
    kd_abs  = _abs(inp["debt_cost"],   INPUT_COL)
    tx_abs  = _abs(inp["tax_rate"],    INPUT_COL)
    # WACC stored as percentage value (e.g. 9.2 meaning 9.2%)
    wacc_formula_pct = (
        f"=((1-{dw_abs}/100)*({rf_abs}/100+{b_abs}*{erp_abs}/100)"
        f"+{dw_abs}/100*{kd_abs}/100*(1-{tx_abs}/100))*100"
    )
    inp["wacc"] = _inp_row("WACC (auto-calculated)", wacc_formula_pct,
                            "CAPM equity cost + after-tax debt cost, auto-calculated", is_wacc=True)

    spacer(ws, r); r += 1

    # ── SECTION 2 + 3: Combined Historical + Projections ─────────────────────
    hist_year_labels = _year_labels(inc, N_HIST)
    proj_year_labels = [f"Proj. {i+1}" for i in range(N_PROJ)]

    sec_hdr(ws, r, "Historical Data + DCF Projections", C1, CE); r += 1

    # Year header row
    ws.row_dimensions[r].height = 16
    wc(ws, r, C1, "", font=_f(9, True, WHITE), bg=NAVY_MED, align=AL_C, border=BORDER_ALL)
    for i, lbl in enumerate(hist_year_labels):
        wc(ws, r, HIST_START + i, lbl,
           font=_f(9, True, WHITE), bg=NAVY_MED, align=AL_C, border=BORDER_ALL)
    for i, lbl in enumerate(proj_year_labels):
        wc(ws, r, PROJ_START + i, lbl,
           font=_f(9, True, WHITE), bg=BLUE_ACC, align=AL_C, border=BORDER_ALL)
    r += 1

    # Absolute refs to inputs (used in projection formulas)
    cagr_abs = _abs(inp["cagr"],         INPUT_COL)
    gm_abs   = _abs(inp["gross_margin"], INPUT_COL)
    em_abs   = _abs(inp["ebitda_margin"],INPUT_COL)
    da_abs   = _abs(inp["da_pct"],       INPUT_COL)
    cx_abs   = _abs(inp["capex_pct"],    INPUT_COL)
    nwc_abs  = _abs(inp["nwc_pct"],      INPUT_COL)
    tx_abs2  = _abs(inp["tax_rate"],     INPUT_COL)
    tg_abs   = _abs(inp["terminal_g"],   INPUT_COL)
    wacc_abs = _abs(inp["wacc"],         INPUT_COL)  # stored as pct

    def _combined_row(label, hist_vals, proj_formulas, alt=False, bold=False):
        """Write one row: hist data in C-G (numeric, $B), proj formulas in H-L."""
        nonlocal r
        bg_h = GRAY_LIGHT if alt else WHITE
        bg_p = BLUE_TINT
        ws.row_dimensions[r].height = 16
        wc(ws, r, C1, label, font=_f(9, bold), bg=BLUE_TINT, align=AL_LI, border=BORDER_ALL)
        for i, v in enumerate(hist_vals):
            cell = ws.cell(row=r, column=HIST_START + i)
            cell.value = v / 1e9 if v is not None else None  # store numeric billions
            cell.font = _f(9); cell.fill = fill(bg_h)
            cell.alignment = AL_R; cell.border = BORDER_ALL
            cell.number_format = '#,##0.0'
        for i, fml in enumerate(proj_formulas):
            cell = ws.cell(row=r, column=PROJ_START + i, value=fml)
            cell.font = _f(9, bold); cell.fill = fill(bg_p)
            cell.alignment = AL_R; cell.border = BORDER_ALL
            cell.number_format = '#,##0.0'
        this_row = r
        r += 1
        return this_row

    def _margin_row(label, num_row, denom_row):
        """Write a margin row with formulas for all 10 data cols."""
        nonlocal r
        ws.row_dimensions[r].height = 14
        wc(ws, r, C1, f"    {label}", font=_f(8, False, DARK_GRAY), bg=WHITE, align=AL_LI, border=BORDER_ALL)
        for i in range(N_HIST):
            col_i = HIST_START + i
            cell = ws.cell(row=r, column=col_i, value=f"={_a(num_row, col_i)}/{_a(denom_row, col_i)}")
            cell.font = _f(8, False, DARK_GRAY); cell.fill = fill(WHITE)
            cell.alignment = AL_R; cell.border = BORDER_ALL; cell.number_format = "0.0%"
        for i in range(N_PROJ):
            col_i = PROJ_START + i
            cell = ws.cell(row=r, column=col_i, value=f"={_a(num_row, col_i)}/{_a(denom_row, col_i)}")
            cell.font = _f(8, False, DARK_GRAY); cell.fill = fill(GRAY_LIGHT)
            cell.alignment = AL_R; cell.border = BORDER_ALL; cell.number_format = "0.0%"
        r += 1

    # ── Revenue ───────────────────────────────────────────────────────────────
    rev_hist = _reversed_series(inc, "Total Revenue", N_HIST)
    rev_proj = []
    for n in range(N_PROJ):
        if n == 0:
            prev_ref = _a(None, PROJ_START - 1)   # placeholder, fixed below
        else:
            prev_ref = None
        rev_proj.append(None)  # fill after we know the row number

    # First write the row to get its row number
    rev_row_num = r  # save current r
    rows["rev"] = _combined_row("Revenue", rev_hist, [""] * N_PROJ, alt=False, bold=True)

    # Now fill proj formulas referencing the correct row
    for n in range(N_PROJ):
        if n == 0:
            prev_ref = _a(rows["rev"], PROJ_START - 1)   # col G most recent hist
        else:
            prev_ref = _a(rows["rev"], PROJ_START + n - 1)
        fml = f"={prev_ref}*(1+{cagr_abs}/100)"
        cell = ws.cell(row=rows["rev"], column=PROJ_START + n, value=fml)
        cell.font = _f(9, True); cell.fill = fill(BLUE_TINT)
        cell.alignment = AL_R; cell.border = BORDER_ALL

    # ── Gross Profit ──────────────────────────────────────────────────────────
    gp_hist = _reversed_series(inc, "Gross Profit", N_HIST)
    gp_proj = [f"={_a(rows['rev'], PROJ_START+n)}*{gm_abs}/100" for n in range(N_PROJ)]
    rows["gp"] = _combined_row("Gross Profit", gp_hist, gp_proj, alt=True)
    _margin_row("Gross Margin %", rows["gp"], rows["rev"])

    # ── EBITDA ────────────────────────────────────────────────────────────────
    ebitda_hist = _reversed_series(inc, "EBITDA", N_HIST)
    ebitda_proj = [f"={_a(rows['rev'], PROJ_START+n)}*{em_abs}/100" for n in range(N_PROJ)]
    rows["ebitda"] = _combined_row("EBITDA", ebitda_hist, ebitda_proj, alt=False)
    _margin_row("EBITDA Margin %", rows["ebitda"], rows["rev"])

    # ── D&A ───────────────────────────────────────────────────────────────────
    da_hist = _reversed_series(cf, "Depreciation And Amortization", N_HIST)
    da_proj = [f"={_a(rows['rev'], PROJ_START+n)}*{da_abs}/100" for n in range(N_PROJ)]
    rows["da"] = _combined_row("D&A", da_hist, da_proj, alt=True)

    # ── EBIT ──────────────────────────────────────────────────────────────────
    ebit_hist = _reversed_series(inc, "Operating Income", N_HIST)
    ebit_proj = [
        f"={_a(rows['ebitda'], PROJ_START+n)}-{_a(rows['da'], PROJ_START+n)}"
        for n in range(N_PROJ)
    ]
    rows["ebit"] = _combined_row("EBIT", ebit_hist, ebit_proj, alt=False)
    _margin_row("EBIT Margin %", rows["ebit"], rows["rev"])

    # ── Net Income (historical only, no projection formula needed in main table)
    ni_hist = _reversed_series(inc, "Net Income", N_HIST)
    rows["ni"] = _combined_row("Net Income", ni_hist, [""] * N_PROJ, alt=True)

    # ── CapEx ─────────────────────────────────────────────────────────────────
    capex_hist = _reversed_series(cf, "Capital Expenditure", N_HIST)
    capex_proj = [f"=-{_a(rows['rev'], PROJ_START+n)}*{cx_abs}/100" for n in range(N_PROJ)]
    rows["capex"] = _combined_row("CapEx", capex_hist, capex_proj, alt=False)

    # ── NWC ───────────────────────────────────────────────────────────────────
    ca_rev = _reversed_series(bs, "Current Assets", N_HIST)
    cl_rev = _reversed_series(bs, "Current Liabilities", N_HIST)
    nwc_hist = [
        (ca_rev[i] - cl_rev[i]) if (ca_rev[i] is not None and cl_rev[i] is not None) else None
        for i in range(N_HIST)
    ]
    rows["nwc"] = _combined_row("Net Working Capital", nwc_hist, [""] * N_PROJ, alt=True)

    # ── FCF (historical only; projected FCF computed in bridge section) ────────
    fcf_hist = _reversed_series(cf, "Free Cash Flow", N_HIST)
    rows["fcf"] = _combined_row("Free Cash Flow", fcf_hist, [""] * N_PROJ, alt=False, bold=True)

    spacer(ws, r); r += 1

    # ── SECTION 4: DCF Bridge ─────────────────────────────────────────────────
    sec_hdr(ws, r, "Detailed Projections - FCF & Present Value", C1, CE); r += 1

    ws.row_dimensions[r].height = 16
    wc(ws, r, C1, "", font=_f(9, True, WHITE), bg=NAVY_MED, align=AL_C, border=BORDER_ALL)
    for i in range(N_HIST):
        wc(ws, r, HIST_START + i, "", font=_f(9, True, WHITE), bg=NAVY_MED, align=AL_C, border=BORDER_ALL)
    for i, lbl in enumerate(proj_year_labels):
        wc(ws, r, PROJ_START + i, lbl, font=_f(9, True, WHITE), bg=BLUE_ACC, align=AL_C, border=BORDER_ALL)
    r += 1

    def _proj_only_row(label, proj_formulas, alt=False, bold=False, fmt='#,##0.0'):
        """Write a row with empty hist cols and projection formulas only."""
        nonlocal r
        bg_p = GRAY_LIGHT if alt else WHITE
        ws.row_dimensions[r].height = 16
        wc(ws, r, C1, label, font=_f(9, bold), bg=BLUE_TINT, align=AL_LI, border=BORDER_ALL)
        for i in range(N_HIST):
            wc(ws, r, HIST_START + i, "", font=_f(9), bg=GRAY_LIGHT, align=AL_C, border=BORDER_ALL)
        for i, fml in enumerate(proj_formulas):
            cell = ws.cell(row=r, column=PROJ_START + i, value=fml)
            cell.font = _f(9, bold); cell.fill = fill(bg_p)
            cell.alignment = AL_R; cell.border = BORDER_ALL
            if fmt: cell.number_format = fmt
        this_row = r
        r += 1
        return this_row

    DCF_FMT = '#,##0.0'   # billions, no embedded quotes

    # Taxes (negative)
    taxes_proj = [
        f"=-MAX({_a(rows['ebit'], PROJ_START+n)},0)*{tx_abs2}/100"
        for n in range(N_PROJ)
    ]
    rows["taxes"] = _proj_only_row("Taxes (–)", taxes_proj, alt=True)

    # NOPAT
    nopat_proj = [
        f"={_a(rows['ebit'], PROJ_START+n)}*(1-{tx_abs2}/100)"
        for n in range(N_PROJ)
    ]
    rows["nopat"] = _proj_only_row("NOPAT", nopat_proj, alt=False, bold=True)

    # D&A addback
    da_add_proj = [f"={_a(rows['da'], PROJ_START+n)}" for n in range(N_PROJ)]
    rows["da_addback"] = _proj_only_row("(+) D&A", da_add_proj, alt=True)

    # CapEx (already negative in rows["capex"])
    capex_add_proj = [f"={_a(rows['capex'], PROJ_START+n)}" for n in range(N_PROJ)]
    rows["capex_bridge"] = _proj_only_row("(–) CapEx", capex_add_proj, alt=False)

    # ΔNWC
    dnwc_proj = []
    for n in range(N_PROJ):
        curr_rev = _a(rows["rev"], PROJ_START + n)
        prev_rev = _a(rows["rev"], PROJ_START + n - 1) if n > 0 else _a(rows["rev"], PROJ_START - 1)
        dnwc_proj.append(f"=-({curr_rev}-{prev_rev})*{nwc_abs}/100")
    rows["dnwc"] = _proj_only_row("(–) ΔNWC", dnwc_proj, alt=True)

    # Projected FCF
    proj_fcf_formulas = [
        f"={_a(rows['nopat'], PROJ_START+n)}"
        f"+{_a(rows['da_addback'], PROJ_START+n)}"
        f"+{_a(rows['capex_bridge'], PROJ_START+n)}"
        f"+{_a(rows['dnwc'], PROJ_START+n)}"
        for n in range(N_PROJ)
    ]
    rows["proj_fcf"] = _proj_only_row("= Projected FCF", proj_fcf_formulas, alt=False, bold=True)

    # Discount Factor
    disc_proj = [f"=1/(1+{wacc_abs}/100)^{n+1}" for n in range(N_PROJ)]
    rows["disc_factor"] = _proj_only_row("Discount Factor", disc_proj, alt=True, fmt="0.0000")

    # PV of FCF
    pv_fcf_proj = [
        f"={_a(rows['proj_fcf'], PROJ_START+n)}*{_a(rows['disc_factor'], PROJ_START+n)}"
        for n in range(N_PROJ)
    ]
    rows["pv_fcf"] = _proj_only_row("PV of FCF", pv_fcf_proj, alt=False, bold=True)

    spacer(ws, r); r += 1

    # ── SECTION 5: Enterprise Value Bridge ───────────────────────────────────
    sec_hdr(ws, r, "Enterprise Value Bridge", C1, CE); r += 1

    cash_fixed   = (cash0  or 0) / 1e9
    debt_fixed   = (debt0  or 0) / 1e9
    shares_fixed = (shares or 1) / 1e9

    def _bridge_line(label, value_formula, highlight=False, big=False, fmt='#,##0.0'):
        nonlocal r
        bg_lbl = GREEN   if big else (NAVY_MED if highlight else BLUE_TINT)
        bg_val = GREEN   if big else (NAVY_MED if highlight else WHITE)
        ft_c   = WHITE   if (big or highlight) else BLACK
        ft_sz  = 13      if big else (10 if highlight else 9)
        ws.row_dimensions[r].height = 26 if big else (18 if highlight else 17)
        wc(ws, r, C1, label,
           font=_f(ft_sz, True, ft_c), bg=bg_lbl, align=AL_LI, border=BORDER_ALL)
        cell = ws.cell(row=r, column=C1 + 1, value=value_formula)
        cell.font = _f(ft_sz, True, ft_c); cell.fill = fill(bg_val)
        cell.alignment = AL_R; cell.border = BORDER_ALL
        cell.number_format = fmt
        merge(ws, r, C1 + 1, r, CE)
        this_row = r
        r += 1
        return this_row

    # Sum PV(FCFs)
    pv_range = f"{_a(rows['pv_fcf'], PROJ_START)}:{_a(rows['pv_fcf'], PROJ_START+N_PROJ-1)}"
    rows["sum_pv"] = _bridge_line("Sum PV(FCFs)", f"=SUM({pv_range})")

    # Terminal FCF
    rows["term_fcf"] = _bridge_line(
        "Terminal FCF (Year 6)",
        f"={_a(rows['proj_fcf'], PROJ_START+N_PROJ-1)}*(1+{tg_abs}/100)"
    )

    # Terminal Value
    rows["tv"] = _bridge_line(
        "Terminal Value (TV)",
        f"={_a(rows['term_fcf'], C1+1)}/({wacc_abs}/100-{tg_abs}/100)"
    )

    # PV(TV)
    rows["pv_tv"] = _bridge_line(
        "PV(TV)",
        f"={_a(rows['tv'], C1+1)}/(1+{wacc_abs}/100)^{N_PROJ}"
    )

    # EV
    rows["ev"] = _bridge_line(
        "→ Implied EV",
        f"={_a(rows['sum_pv'], C1+1)}+{_a(rows['pv_tv'], C1+1)}",
        highlight=True
    )

    # Cash & Debt (fixed values in $B)
    rows["cash_b"] = _bridge_line("(+) Cash ($B)",       cash_fixed)
    rows["debt_b"] = _bridge_line("(–) Total Debt ($B)", -debt_fixed)

    # Equity Value
    rows["equity_val"] = _bridge_line(
        "= Equity Value ($B)",
        f"={_a(rows['ev'], C1+1)}+{_a(rows['cash_b'], C1+1)}+{_a(rows['debt_b'], C1+1)}",
        highlight=True
    )

    # Shares
    rows["shares_b"] = _bridge_line("Shares Outstanding ($B)", shares_fixed, fmt='#,##0.0')

    # Fair value per share (BIG row)
    rows["fair_value"] = _bridge_line(
        "→ FAIR VALUE PER SHARE",
        f"={_a(rows['equity_val'], C1+1)}/{_a(rows['shares_b'], C1+1)}",
        big=True
    )
    # Set number format on fair value cell
    ws.cell(row=rows["fair_value"], column=C1 + 1).number_format = '$#,##0.00'

    # Market price
    rows["market_price"] = _bridge_line("Market Price", price or 0)
    ws.cell(row=rows["market_price"], column=C1 + 1).number_format = '$#,##0.00'

    # Upside
    ws.row_dimensions[r].height = 17
    wc(ws, r, C1, "Implied Upside",
       font=_f(9, True), bg=BLUE_TINT, align=AL_LI, border=BORDER_ALL)
    upside_fml = f"={_a(rows['fair_value'], C1+1)}/{_a(rows['market_price'], C1+1)}-1"
    cell = ws.cell(row=r, column=C1 + 1, value=upside_fml)
    cell.font = _f(9, True); cell.fill = fill(BLUE_TINT)
    cell.alignment = AL_C; cell.border = BORDER_ALL; cell.number_format = "0.0%"
    merge(ws, r, C1 + 1, r, CE)
    r += 1

    # TV as % of EV
    ws.row_dimensions[r].height = 17
    wc(ws, r, C1, "TV as % of EV",
       font=_f(9, True), bg=BLUE_TINT, align=AL_LI, border=BORDER_ALL)
    tv_pct_fml = f"={_a(rows['pv_tv'], C1+1)}/({_a(rows['sum_pv'], C1+1)}+{_a(rows['pv_tv'], C1+1)})"
    cell = ws.cell(row=r, column=C1 + 1, value=tv_pct_fml)
    cell.font = _f(9); cell.fill = fill(WHITE)
    cell.alignment = AL_C; cell.border = BORDER_ALL; cell.number_format = "0.0%"
    merge(ws, r, C1 + 1, r, CE)
    r += 1

    spacer(ws, r); r += 1

    # Footer
    ws.row_dimensions[r].height = 20
    wc(ws, r, C1,
       "  Yellow cells are editable. Adjust WACC inputs, margins, or CAGR to see the impact on fair value per share. "
       "WACC is auto-calculated from beta, risk-free rate, and capital structure.",
       font=_f(8, False, MID_GRAY), bg=GRAY_LIGHT, align=AL_W)
    merge(ws, r, C1, r, CE)

    return ws
