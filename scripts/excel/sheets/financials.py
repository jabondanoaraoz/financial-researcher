"""
Financials Sheet  (v3)
Company profile, income statement (5y), cash flow (5y), balance sheet (5y).
Oldest year LEFT, most recent year RIGHT.
Margin column (col H) uses Excel formulas referencing the data cells.

Author: Joaquin Abondano w/ Claude Code
"""

import math
from openpyxl.utils import get_column_letter as _gcl

from ..styles import (
    _f, fill, merge, wc, sec_hdr, col_hdr, spacer, hide_gridlines,
    NAVY_DARK, NAVY_MED, BLUE_ACC, GREEN, AMBER, RED,
    BLUE_TINT, GRAY_LIGHT, WHITE, BLACK, DARK_GRAY, MID_GRAY,
    SIGNAL_BG, SIGNAL_TEXT, ACTION_TEXT,
    AL_L, AL_LI, AL_C, AL_R, AL_W,
    BORDER_ALL,
)

_COLS = {
    "A": 2,   # spacer
    "B": 26,  # label
    "C": 12,  # FY oldest
    "D": 12,
    "E": 12,
    "F": 12,
    "G": 12,  # FY most recent
    "H": 10,  # margin / ratio col
    "I": 2,   # spacer
}
C1 = 2    # col B
CE = 8    # col H
N_YEARS = 5

INPUT_YELLOW = "FFFF99"
DATA_COL_START = C1 + 1   # col C = column index 3
DATA_COL_END   = C1 + N_YEARS  # col G = column index 7 (most recent year)


def _a(row, col, ar=False, ac=False):
    """Excel cell address string, optionally absolute rows/cols."""
    c = ("$" if ac else "") + _gcl(col)
    r = ("$" if ar else "") + str(row)
    return c + r


# ── Formatters ────────────────────────────────────────────────────────────────

def _b(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return "—"
    return f"${v / 1e9:.1f}B"

def _pct(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return "—"
    return f"{v:.1f}%"

def _mult(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return "—"
    return f"{v:.1f}x"

def _num(v, d=2):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return "—"
    return f"{v:.{d}f}"

def _price(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return "—"
    return f"${v:,.2f}"

def _mcap(v):
    if v is None:
        return "—"
    if v >= 1e12:
        return f"${v/1e12:.2f}T"
    return f"${v/1e9:.1f}B"


# ── DataFrame helpers ─────────────────────────────────────────────────────────

def _get(df, row_key, col_idx=0):
    if df is None or df.empty or row_key not in df.index:
        return None
    try:
        v = float(df.iloc[df.index.get_loc(row_key), col_idx])
        return None if math.isnan(v) else v
    except (TypeError, ValueError, IndexError):
        return None


def _reversed_series(df, row_key, n=5):
    """Return up to n values oldest-first (reversed from DataFrame newest-first)."""
    if df is None or df.empty or row_key not in df.index:
        return [None] * n
    vals = []
    for v in df.loc[row_key].values[:n]:
        try:
            fv = float(v)
            vals.append(None if math.isnan(fv) else fv)
        except (TypeError, ValueError):
            vals.append(None)
    while len(vals) < n:
        vals.append(None)
    vals.reverse()
    return vals


def _reversed_labels(df, n=5):
    """Return up to n FY labels oldest-first."""
    if df is None or df.empty:
        return ["—"] * n
    labels = []
    for col in list(df.columns)[:n]:
        try:
            labels.append(f"FY{col.year}")
        except AttributeError:
            labels.append(str(col)[:6])
    while len(labels) < n:
        labels.insert(0, "—")
    labels.reverse()
    return labels


# ── Sheet builder ─────────────────────────────────────────────────────────────

def build(wb, result):
    ws = wb.create_sheet("Financials")
    hide_gridlines(ws)
    ws.sheet_properties.tabColor = NAVY_MED

    for col, w in _COLS.items():
        ws.column_dimensions[col].width = w

    ticker = result.get("ticker", "")
    info   = (result.get("company_data") or {}).get("info") or {}
    km     = (result.get("company_data") or {}).get("key_metrics") or {}
    fins   = result.get("financials") or {}

    inc = fins.get("income_statement")
    cf  = fins.get("cash_flow")
    bs  = fins.get("balance_sheet")

    year_labels = _reversed_labels(inc, N_YEARS)

    # Track row numbers for formula back-references
    rows = {}

    r = 1
    spacer(ws, r, 8); r += 1

    # ── Company Header ────────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 30
    name = info.get("name", ticker)
    wc(ws, r, C1, f"  {name}  ({ticker})",
       font=_f(14, True, WHITE), bg=NAVY_DARK, align=AL_L)
    merge(ws, r, C1, r, CE)
    r += 1

    ws.row_dimensions[r].height = 18
    sector   = info.get("sector",   "—")
    industry = info.get("industry", "—")
    exchange = info.get("exchange", "—")
    country  = info.get("country",  "—")
    wc(ws, r, C1, f"  {sector}  |  {industry}  |  {exchange}  |  {country}",
       font=_f(9, False, MID_GRAY), bg=GRAY_LIGHT, align=AL_L)
    merge(ws, r, C1, r, CE)
    r += 1

    employees = info.get("employees")
    emp_s = f"{employees:,}" if employees else "—"
    website = info.get("website", "")
    wc(ws, r, C1, f"  Employees: {emp_s}   |   {website}",
       font=_f(8, False, DARK_GRAY), bg=GRAY_LIGHT, align=AL_L)
    merge(ws, r, C1, r, CE)
    r += 1
    spacer(ws, r, 4); r += 1

    # ── Business Description ──────────────────────────────────────────────────
    desc = info.get("description", "")
    if desc:
        sec_hdr(ws, r, "Business Description", C1, CE); r += 1
        ws.row_dimensions[r].height = 60
        wc(ws, r, C1, desc,
           font=_f(8.5, False, BLACK), bg=WHITE, align=AL_W, border=BORDER_ALL)
        merge(ws, r, C1, r, CE)
        r += 1
        spacer(ws, r, 4); r += 1

    # ── Key Metrics ───────────────────────────────────────────────────────────
    sec_hdr(ws, r, "Key Metrics", C1, CE); r += 1

    price   = km.get("current_price")
    mkt_cap = info.get("market_cap") or km.get("market_cap")
    ev      = km.get("enterprise_value")

    metrics_grid = [
        [("Current Price",    _price(price)),
         ("Market Cap",       _mcap(mkt_cap)),
         ("Enterprise Value", _mcap(ev))],
        [("P/E Ratio",        _mult(km.get("pe_ratio"))),
         ("Forward P/E",      _mult(km.get("forward_pe"))),
         ("PEG Ratio",        _num(km.get("peg_ratio"), 2))],
        [("P/B Ratio",        _mult(km.get("pb_ratio"))),
         ("P/S Ratio",        _mult(km.get("ps_ratio"))),
         ("EV / EBITDA",      _mult(km.get("ev_ebitda")))],
        [("Beta",             _num(km.get("beta"), 3)),
         ("52W High",         _price(km.get("52w_high"))),
         ("52W Low",          _price(km.get("52w_low")))],
        [("Dividend Yield",   _pct(km.get("dividend_yield"))),
         ("Short % Float",    _pct(km.get("short_percent_of_float"))),
         ("Shares Out.",      _mcap(km.get("shares_outstanding")))],
    ]

    for row_data in metrics_grid:
        ws.row_dimensions[r].height = 17
        for i, (lbl, val) in enumerate(row_data):
            c_lbl = C1 + i * 2
            c_val = c_lbl + 1
            wc(ws, r, c_lbl, lbl, font=_f(8.5, True), bg=BLUE_TINT, align=AL_L, border=BORDER_ALL)
            wc(ws, r, c_val, val, font=_f(9, False),   bg=WHITE,     align=AL_R, border=BORDER_ALL)
        r += 1

    spacer(ws, r); r += 1

    # ── Helper to write one financial row ──────────────────────────────────────
    def _fin_row(label, row_key, df_src, alt=False):
        """Write a financial row with 5 numeric data cols (in $B). Returns row number."""
        nonlocal r
        bg = GRAY_LIGHT if alt else WHITE
        ws.row_dimensions[r].height = 16
        wc(ws, r, C1, label, font=_f(9, True), bg=BLUE_TINT, align=AL_LI, border=BORDER_ALL)
        vals = _reversed_series(df_src, row_key, N_YEARS)
        for i, v in enumerate(vals):
            cell = ws.cell(row=r, column=C1 + 1 + i)
            cell.value = v / 1e9 if v is not None else None  # store numeric billions
            cell.font = _f(9); cell.fill = fill(bg)
            cell.alignment = AL_R; cell.border = BORDER_ALL
            cell.number_format = '#,##0.0'
        # Margin col placeholder (H)
        wc(ws, r, CE, "—", font=_f(9), bg=bg, align=AL_C, border=BORDER_ALL)
        this_row = r
        r += 1
        return this_row

    def _set_margin_formula(data_row, ref_row):
        """Overwrite col H of data_row with =G{data_row}/G{ref_row}."""
        formula = f"={_a(data_row, DATA_COL_END)}/{_a(ref_row, DATA_COL_END)}"
        cell = ws.cell(row=data_row, column=CE, value=formula)
        cell.font       = _f(9, True)
        cell.fill       = fill(BLUE_TINT)
        cell.alignment  = AL_C
        cell.border     = BORDER_ALL
        cell.number_format = "0.0%"

    # ── Income Statement ──────────────────────────────────────────────────────
    sec_hdr(ws, r, "Income Statement  (annual, $B)", C1, CE); r += 1
    col_hdr(ws, r, ["Metric"] + year_labels + ["Margin (MRY)"], C1)
    r += 1

    rows["revenue"]     = _fin_row("Revenue",          "Total Revenue",    inc, alt=False)
    rows["gross_profit"]= _fin_row("Gross Profit",      "Gross Profit",     inc, alt=True)
    rows["ebitda"]      = _fin_row("EBITDA",            "EBITDA",           inc, alt=False)
    rows["ebit"]        = _fin_row("Operating Income",  "Operating Income", inc, alt=True)
    rows["net_income"]  = _fin_row("Net Income",        "Net Income",       inc, alt=False)

    # Set margin formulas (all reference revenue row most-recent col G)
    _set_margin_formula(rows["gross_profit"], rows["revenue"])
    _set_margin_formula(rows["ebitda"],       rows["revenue"])
    _set_margin_formula(rows["ebit"],         rows["revenue"])
    _set_margin_formula(rows["net_income"],   rows["revenue"])

    spacer(ws, r); r += 1

    # ── Cash Flow ─────────────────────────────────────────────────────────────
    sec_hdr(ws, r, "Cash Flow  (annual, $B)", C1, CE); r += 1
    col_hdr(ws, r, ["Metric"] + year_labels + ["FCF Margin"], C1)
    r += 1

    rows["ocf"]   = _fin_row("Operating Cash Flow", "Operating Cash Flow", cf, alt=False)
    rows["capex"] = _fin_row("Capital Expenditure", "Capital Expenditure", cf, alt=True)
    rows["fcf"]   = _fin_row("Free Cash Flow",       "Free Cash Flow",     cf, alt=False)

    # FCF margin formula: =G{fcf}/G{rev}
    _set_margin_formula(rows["fcf"], rows["revenue"])

    spacer(ws, r); r += 1

    # ── Balance Sheet ─────────────────────────────────────────────────────────
    sec_hdr(ws, r, "Balance Sheet Evolution  (annual, $B)", C1, CE); r += 1
    col_hdr(ws, r, ["Metric"] + year_labels + ["Ratio (MRY)"], C1)
    r += 1

    rows["cash"]         = _fin_row("Cash & Equivalents",
                                     "Cash Cash Equivalents And Short Term Investments",
                                     bs, alt=False)
    rows["debt"]         = _fin_row("Total Debt",         "Total Debt",         bs, alt=True)
    rows["equity"]       = _fin_row("Stockholders Equity","Stockholders Equity",bs, alt=False)
    rows["current_assets"]    = _fin_row("Current Assets",     "Current Assets",     bs, alt=True)
    rows["current_liabilities"]= _fin_row("Current Liabilities","Current Liabilities",bs, alt=False)

    # Current Ratio row — formula only in H col; fill "—" in data cols
    ws.row_dimensions[r].height = 16
    wc(ws, r, C1, "Current Ratio", font=_f(9, True), bg=BLUE_TINT, align=AL_LI, border=BORDER_ALL)
    for i in range(N_YEARS):
        wc(ws, r, C1 + 1 + i, "—", font=_f(9), bg=GRAY_LIGHT, align=AL_R, border=BORDER_ALL)
    cr_formula = f"={_a(rows['current_assets'], DATA_COL_END)}/{_a(rows['current_liabilities'], DATA_COL_END)}"
    cell = ws.cell(row=r, column=CE, value=cr_formula)
    cell.font = _f(9, True); cell.fill = fill(BLUE_TINT)
    cell.alignment = AL_C; cell.border = BORDER_ALL; cell.number_format = "0.00"
    rows["current_ratio"] = r
    r += 1

    # Net Debt — each year col formula: =debt_cell - cash_cell
    ws.row_dimensions[r].height = 16
    wc(ws, r, C1, "Net Debt", font=_f(9, True, WHITE), bg=NAVY_MED, align=AL_LI, border=BORDER_ALL)
    for i in range(N_YEARS):
        col_i = C1 + 1 + i
        nd_formula = f"={_a(rows['debt'], col_i)}-{_a(rows['cash'], col_i)}"
        cell = ws.cell(row=r, column=col_i, value=nd_formula)
        cell.font = _f(9, False, WHITE); cell.fill = fill(NAVY_MED)
        cell.alignment = AL_R; cell.border = BORDER_ALL
        cell.number_format = '#,##0.0'
    wc(ws, r, CE, "—", font=_f(9, False, WHITE), bg=NAVY_MED, align=AL_C, border=BORDER_ALL)
    rows["net_debt"] = r
    r += 1

    spacer(ws, r); r += 1

    # Footer
    ws.row_dimensions[r].height = 14
    wc(ws, r, C1,
       "  Source: yfinance / Alpha Vantage. Data ordered oldest → newest (left → right). "
       "Margin formulas in col H reference most recent year (col G). "
       "Empty cells (—) indicate data not reported or not available from the API for this company.",
       font=_f(8, False, MID_GRAY), bg=GRAY_LIGHT, align=AL_L)
    merge(ws, r, C1, r, CE)

    return ws
