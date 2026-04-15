"""
Multiples Sheet  (v3 — NEW)
Peer-based valuation tool. User sets peer weights and method weights.
All derived cells use Excel formulas. Yellow cells are editable.

Author: Joaquin Abondano w/ Claude Code
"""

import math
from openpyxl.utils import get_column_letter as _gcl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

from ..styles import (
    _f, fill, merge, wc, sec_hdr, col_hdr, spacer, hide_gridlines,
    NAVY_DARK, NAVY_MED, BLUE_ACC, GREEN, AMBER, RED,
    BLUE_TINT, GRAY_LIGHT, WHITE, BLACK, DARK_GRAY, MID_GRAY,
    SIGNAL_BG,
    AL_L, AL_LI, AL_C, AL_R, AL_W, AL_WC,
    BORDER_ALL,
)

_COLS = {
    "A": 2,   # spacer
    "B": 18,  # label / empresa
    "C": 10,  # peso (%)
    "D": 11,  # P/S
    "E": 11,  # P/E
    "F": 11,  # Fwd P/E
    "G": 11,  # EV/EBITDA
    "H": 11,  # P/FCF
    "I": 11,  # extra (upside / contribucion)
    "J": 2,   # spacer
}
C1 = 2    # col B
CE = 9    # col I

INPUT_YELLOW = "FFFF99"

ACTION_BG = {
    "buy":   GREEN,
    "hold":  AMBER,
    "sell":  RED,
    "short": RED,
    "cover": GREEN,
}


def _a(row, col, ar=False, ac=False):
    c = ("$" if ac else "") + _gcl(col)
    r = ("$" if ar else "") + str(row)
    return c + r


def _b(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    return v / 1e9

def _safe(v, default=None):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return default
    return v

def _fmt_b(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return "—"
    return f"${v/1e9:.1f}B"

def _fmt_mult(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    return v


def build(wb, result):
    ws = wb.create_sheet("Multiples")
    hide_gridlines(ws)
    ws.sheet_properties.tabColor = BLUE_ACC

    for col, w in _COLS.items():
        ws.column_dimensions[col].width = w

    ticker  = result.get("ticker", "")
    info    = (result.get("company_data") or {}).get("info") or {}
    km      = (result.get("company_data") or {}).get("key_metrics") or {}
    fins    = result.get("financials") or {}
    peers   = result.get("peers_data") or {}

    inc = fins.get("income_statement")
    cf  = fins.get("cash_flow")
    bs  = fins.get("balance_sheet")

    def _get(df, row_key, col_idx=0):
        if df is None or df.empty or row_key not in df.index:
            return None
        try:
            v = float(df.iloc[df.index.get_loc(row_key), col_idx])
            return None if math.isnan(v) else v
        except (TypeError, ValueError, IndexError):
            return None

    rows = {}   # track cell addresses for formulas

    r = 1
    spacer(ws, r, 8); r += 1

    # Title
    ws.row_dimensions[r].height = 28
    wc(ws, r, C1, f"  PEER MULTIPLES VALUATION  —  {ticker}",
       font=_f(14, True, WHITE), bg=NAVY_DARK, align=AL_L)
    merge(ws, r, C1, r, CE)
    r += 1
    spacer(ws, r, 4); r += 1

    # ── SECTION 1: Comparable Companies ──────────────────────────────────────
    sec_hdr(ws, r, "1. Comparable Companies", C1, CE); r += 1

    col_hdr(ws, r, ["Company", "Weight (%)", "P/S", "P/E", "Fwd P/E", "EV/EBITDA", "P/FCF", "Beta"],
            C1)
    r += 1

    n_peers = len(peers)
    equal_weight = round(100.0 / n_peers, 1) if n_peers > 0 else 25.0

    peer_start_row = r
    peer_rows = []

    for i, (pt, pd_) in enumerate(peers.items()):
        alt = (i % 2 == 0)
        bg  = GRAY_LIGHT if alt else WHITE
        ws.row_dimensions[r].height = 17

        # Empresa name
        wc(ws, r, C1, pt, font=_f(9, True), bg=BLUE_TINT, align=AL_C, border=BORDER_ALL)

        # Peso — INPUT_YELLOW, pre-filled with equal weight
        cell = ws.cell(row=r, column=C1 + 1, value=equal_weight)
        cell.font = _f(9, True); cell.fill = fill(INPUT_YELLOW)
        cell.alignment = AL_C; cell.border = BORDER_ALL
        cell.number_format = "0.0"

        # Multiples: P/S, P/E, Fwd P/E, EV/EBITDA, P/FCF, Beta
        mult_keys = ["ps_ratio", "pe_ratio", "forward_pe", "ev_ebitda", "price_to_fcf", "beta"]
        for j, key in enumerate(mult_keys):
            v = _safe(pd_.get(key))
            wc(ws, r, C1 + 2 + j, v, font=_f(9), bg=bg, align=AL_R, border=BORDER_ALL)

        peer_rows.append(r)
        r += 1

    peer_end_row = r - 1

    # Suma pesos row
    ws.row_dimensions[r].height = 15
    wc(ws, r, C1, "Weight sum:", font=_f(8, True, DARK_GRAY), bg=GRAY_LIGHT, align=AL_R, border=BORDER_ALL)
    suma_formula = f"=SUM({_a(peer_start_row, C1+1)}:{_a(peer_end_row, C1+1)})"
    cell = ws.cell(row=r, column=C1 + 1, value=suma_formula)
    cell.font = _f(8, True, DARK_GRAY); cell.fill = fill(GRAY_LIGHT)
    cell.alignment = AL_C; cell.border = BORDER_ALL; cell.number_format = "0.0"
    merge(ws, r, C1, r, C1)
    rows["suma_pesos"] = r
    # Red conditional formatting if peer weights exceed 100%
    _red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    ws.conditional_formatting.add(
        f"{_gcl(C1+1)}{r}",
        CellIsRule(operator="greaterThan", formula=["100"], fill=_red_fill)
    )
    r += 1

    # Promedio Ponderado row
    ws.row_dimensions[r].height = 18
    wc(ws, r, C1, "Weighted Average", font=_f(10, True, WHITE), bg=NAVY_MED, align=AL_L, border=BORDER_ALL)
    wc(ws, r, C1 + 1, "—", font=_f(9, False, WHITE), bg=NAVY_MED, align=AL_C, border=BORDER_ALL)

    # Weighted avg for each multiple col (D through I = C1+2 through C1+7)
    wtd_rows = {}
    mult_col_names = ["ps", "pe", "fwdpe", "evebitda", "pfcf", "beta"]
    for j, name in enumerate(mult_col_names):
        mult_col = C1 + 2 + j
        weight_range = f"{_a(peer_start_row, C1+1)}:{_a(peer_end_row, C1+1)}"
        data_range   = f"{_a(peer_start_row, mult_col)}:{_a(peer_end_row, mult_col)}"
        wtd_formula  = f"=SUMPRODUCT({weight_range},{data_range})/SUM({weight_range})"
        cell = ws.cell(row=r, column=mult_col, value=wtd_formula)
        cell.font = _f(9, True, WHITE); cell.fill = fill(NAVY_MED)
        cell.alignment = AL_C; cell.border = BORDER_ALL; cell.number_format = "0.0"
        wtd_rows[name] = r  # same row, different cols

    rows["wtd_row"] = r
    rows["wtd_ps_col"]      = C1 + 2
    rows["wtd_pe_col"]      = C1 + 3
    rows["wtd_fwdpe_col"]   = C1 + 4
    rows["wtd_evebitda_col"]= C1 + 5
    rows["wtd_pfcf_col"]    = C1 + 6
    r += 1

    spacer(ws, r); r += 1

    # ── SECTION 2: Datos Financieros de la Empresa ────────────────────────────
    sec_hdr(ws, r, "2. Company Financials  (most recent year)", C1, CE); r += 1

    # Get most recent financials (col index 0 = most recent in API format)
    rev_val    = _get(inc, "Total Revenue", 0)
    gp_val     = _get(inc, "Gross Profit",  0)
    ebitda_val = _get(inc, "EBITDA",        0)
    ni_val     = _get(inc, "Net Income",    0)
    ocf_val    = _get(cf,  "Operating Cash Flow",    0)
    capex_val  = _get(cf,  "Capital Expenditure",    0)
    fcf_raw    = _get(cf,  "Free Cash Flow", 0)
    fcf_val    = fcf_raw if fcf_raw is not None else (
        (ocf_val + capex_val) if (ocf_val and capex_val) else None
    )
    shares_val = km.get("shares_outstanding")
    cash_val   = _get(bs, "Cash Cash Equivalents And Short Term Investments", 0)
    debt_val   = _get(bs, "Total Debt", 0)
    price_val  = km.get("current_price")

    fin_data = [
        ("Revenue (MRY)",          rev_val,    "rev"),
        ("Gross Profit (MRY)",     gp_val,     "gp"),
        ("EBITDA (MRY)",           ebitda_val, "ebitda"),
        ("Net Income (MRY)",       ni_val,     "ni"),
        ("Free Cash Flow (MRY)",   fcf_val,    "fcf"),
        ("Shares Outstanding",     shares_val, "shares"),
        ("Cash & Equivalents",     cash_val,   "cash"),
        ("Total Debt",             debt_val,   "debt"),
    ]

    for i, (lbl, val, key) in enumerate(fin_data):
        alt = (i % 2 == 0)
        bg  = GRAY_LIGHT if alt else WHITE
        ws.row_dimensions[r].height = 16
        wc(ws, r, C1, lbl, font=_f(9, True), bg=BLUE_TINT, align=AL_LI, border=BORDER_ALL)
        # Value cell — store raw numeric
        display_val = val / 1e9 if (val is not None and key not in ("shares",)) else val
        if key == "shares":
            display_val = val / 1e9 if val else None
            fmt = '#,##0.0'
        else:
            fmt = '#,##0.0'
        cell = ws.cell(row=r, column=C1 + 1, value=display_val)
        cell.font = _f(9); cell.fill = fill(bg)
        cell.alignment = AL_R; cell.border = BORDER_ALL
        cell.number_format = fmt
        merge(ws, r, C1 + 1, r, CE)
        rows[f"fin_{key}"] = r
        r += 1

    # Net Debt — formula
    ws.row_dimensions[r].height = 16
    wc(ws, r, C1, "Net Debt", font=_f(9, True), bg=BLUE_TINT, align=AL_LI, border=BORDER_ALL)
    nd_formula = f"={_a(rows['fin_debt'], C1+1)}-{_a(rows['fin_cash'], C1+1)}"
    cell = ws.cell(row=r, column=C1 + 1, value=nd_formula)
    cell.font = _f(9); cell.fill = fill(GRAY_LIGHT)
    cell.alignment = AL_R; cell.border = BORDER_ALL; cell.number_format = '#,##0.0'
    merge(ws, r, C1 + 1, r, CE)
    rows["fin_net_debt"] = r
    r += 1

    # Current Market Price
    ws.row_dimensions[r].height = 16
    wc(ws, r, C1, "Current Market Price", font=_f(9, True), bg=BLUE_TINT, align=AL_LI, border=BORDER_ALL)
    cell = ws.cell(row=r, column=C1 + 1, value=price_val)
    cell.font = _f(9); cell.fill = fill(WHITE)
    cell.alignment = AL_R; cell.border = BORDER_ALL; cell.number_format = '$#,##0.00'
    merge(ws, r, C1 + 1, r, CE)
    rows["fin_price"] = r
    r += 1

    spacer(ws, r); r += 1

    # ── SECTION 3: Valoración por Método ─────────────────────────────────────
    sec_hdr(ws, r, "3. Valuation by Method", C1, CE); r += 1

    col_hdr(ws, r, ["Method", "Base Metric", "Wtd. Multiple", "Implied EV/Cap", "Implied Price", "vs Market", "Upside"],
            C1)
    r += 1

    # Helper: references to the input data cells
    rev_ref    = _a(rows["fin_rev"],    C1 + 1)
    ni_ref     = _a(rows["fin_ni"],     C1 + 1)
    ebitda_ref = _a(rows["fin_ebitda"], C1 + 1)
    fcf_ref    = _a(rows["fin_fcf"],    C1 + 1)
    cash_ref   = _a(rows["fin_cash"],   C1 + 1)
    debt_ref   = _a(rows["fin_debt"],   C1 + 1)
    shares_ref = _a(rows["fin_shares"], C1 + 1)
    price_ref  = _a(rows["fin_price"],  C1 + 1)

    wtd_row = rows["wtd_row"]
    ps_wtd_ref      = _a(wtd_row, rows["wtd_ps_col"])
    pe_wtd_ref      = _a(wtd_row, rows["wtd_pe_col"])
    fwdpe_wtd_ref   = _a(wtd_row, rows["wtd_fwdpe_col"])
    evebitda_wtd_ref= _a(wtd_row, rows["wtd_evebitda_col"])
    pfcf_wtd_ref    = _a(wtd_row, rows["wtd_pfcf_col"])

    # Method rows. Cols: B=method, C=metric base, D=multiple, E=EV/cap, F=implied price, G=vs mkt, H=upside (I unused)
    methods = [
        {
            "name": "P/S",
            "metric": "Revenue",
            "multiple_ref": ps_wtd_ref,
            "type": "mktcap",   # implied_cap = multiple * metric
            "metric_ref": rev_ref,
        },
        {
            "name": "P/E",
            "metric": "Net Income",
            "multiple_ref": pe_wtd_ref,
            "type": "mktcap",
            "metric_ref": ni_ref,
        },
        {
            "name": "Fwd P/E",
            "metric": "Net Income × 1.07",
            "multiple_ref": fwdpe_wtd_ref,
            "type": "mktcap_fwd",
            "metric_ref": ni_ref,
        },
        {
            "name": "EV/EBITDA",
            "metric": "EBITDA",
            "multiple_ref": evebitda_wtd_ref,
            "type": "ev",
            "metric_ref": ebitda_ref,
        },
        {
            "name": "P/FCF",
            "metric": "Free Cash Flow",
            "multiple_ref": pfcf_wtd_ref,
            "type": "mktcap",
            "metric_ref": fcf_ref,
        },
    ]

    rows["method_implied_price"] = {}

    for i, m in enumerate(methods):
        alt = (i % 2 == 0)
        bg  = GRAY_LIGHT if alt else WHITE
        ws.row_dimensions[r].height = 17

        # B: method name
        wc(ws, r, C1,     m["name"],   font=_f(9, True), bg=BLUE_TINT, align=AL_C, border=BORDER_ALL)
        # C: metric base
        wc(ws, r, C1 + 1, m["metric"], font=_f(8, False, DARK_GRAY), bg=bg, align=AL_L, border=BORDER_ALL)
        # D: multiple weighted avg (reference to wtd row)
        ref_cell = ws.cell(row=r, column=C1 + 2, value=f"={m['multiple_ref']}")
        ref_cell.font = _f(9); ref_cell.fill = fill(bg)
        ref_cell.alignment = AL_R; ref_cell.border = BORDER_ALL; ref_cell.number_format = "0.0"

        # E: Implied EV or Market Cap
        if m["type"] == "mktcap":
            ev_formula = f"={m['multiple_ref']}*{m['metric_ref']}"
        elif m["type"] == "mktcap_fwd":
            ev_formula = f"={m['multiple_ref']}*{m['metric_ref']}*1.07"
        elif m["type"] == "ev":
            ev_formula = f"={m['multiple_ref']}*{m['metric_ref']}"
        ev_cell_addr = _a(r, C1 + 3)
        cell = ws.cell(row=r, column=C1 + 3, value=ev_formula)
        cell.font = _f(9); cell.fill = fill(bg)
        cell.alignment = AL_R; cell.border = BORDER_ALL; cell.number_format = '#,##0.0'

        # F: Implied Price
        if m["type"] == "ev":
            # Equity = EV + cash - debt; Price = equity / shares
            equity_formula = f"=({ev_cell_addr}+{cash_ref}-{debt_ref})"
            price_formula  = f"=({ev_cell_addr}+{cash_ref}-{debt_ref})/{shares_ref}"
        else:
            # Market Cap / shares
            price_formula = f"={ev_cell_addr}/{shares_ref}"

        implied_price_cell_addr = _a(r, C1 + 4)
        cell = ws.cell(row=r, column=C1 + 4, value=price_formula)
        cell.font = _f(9, True); cell.fill = fill(bg)
        cell.alignment = AL_R; cell.border = BORDER_ALL; cell.number_format = '$#,##0.00'
        rows["method_implied_price"][m["name"]] = (r, C1 + 4)

        # G: vs Mercado (difference in %)
        vs_formula = f"={implied_price_cell_addr}/{price_ref}-1"
        cell = ws.cell(row=r, column=C1 + 5, value=vs_formula)
        cell.font = _f(9); cell.fill = fill(bg)
        cell.alignment = AL_C; cell.border = BORDER_ALL; cell.number_format = "0.0%"

        # H: Upside (same formula — upside = implied/market - 1)
        up_formula = f"={implied_price_cell_addr}/{price_ref}-1"
        cell = ws.cell(row=r, column=C1 + 6, value=up_formula)
        cell.font = _f(9); cell.fill = fill(bg)
        cell.alignment = AL_C; cell.border = BORDER_ALL; cell.number_format = "0.0%"

        r += 1

    spacer(ws, r); r += 1

    # ── SECTION 4: Precio Objetivo Ponderado ──────────────────────────────────
    sec_hdr(ws, r, "4. Blended Price Target", C1, CE); r += 1

    col_hdr(ws, r, ["Method", "Weight (%)", "Implied Price", "Contribution", "", "", "", ""],
            C1)
    r += 1

    method_names = [m["name"] for m in methods]
    n_methods    = len(method_names)
    equal_mw     = round(100.0 / n_methods, 1) if n_methods > 0 else 20.0

    contrib_cells = []
    wt_method_start = r
    wt_method_rows  = []

    for i, m in enumerate(methods):
        alt = (i % 2 == 0)
        bg  = GRAY_LIGHT if alt else WHITE
        ws.row_dimensions[r].height = 17

        # B: method name
        wc(ws, r, C1, m["name"], font=_f(9, True), bg=BLUE_TINT, align=AL_C, border=BORDER_ALL)

        # C: method weight — INPUT_YELLOW
        cell = ws.cell(row=r, column=C1 + 1, value=equal_mw)
        cell.font = _f(9, True); cell.fill = fill(INPUT_YELLOW)
        cell.alignment = AL_C; cell.border = BORDER_ALL; cell.number_format = "0.0"
        wt_method_rows.append(r)

        # D: Implied Price (reference)
        imp_r, imp_c = rows["method_implied_price"][m["name"]]
        imp_ref = _a(imp_r, imp_c)
        cell = ws.cell(row=r, column=C1 + 2, value=f"={imp_ref}")
        cell.font = _f(9); cell.fill = fill(bg)
        cell.alignment = AL_R; cell.border = BORDER_ALL; cell.number_format = '$#,##0.00'

        # E: Contribution = weight * implied_price / 100
        contrib_formula = f"={_a(r, C1+1)}/100*{_a(r, C1+2)}"
        cell = ws.cell(row=r, column=C1 + 3, value=contrib_formula)
        cell.font = _f(9); cell.fill = fill(bg)
        cell.alignment = AL_R; cell.border = BORDER_ALL; cell.number_format = '$#,##0.00'
        contrib_cells.append(_a(r, C1 + 3))

        # fill remaining cols with empty
        for jj in range(4, 8):
            wc(ws, r, C1 + jj, "", font=_f(9), bg=bg, align=AL_C, border=BORDER_ALL)

        r += 1

    wt_method_end = r - 1

    # Suma pesos métodos
    ws.row_dimensions[r].height = 14
    wc(ws, r, C1, "Weight sum:", font=_f(8, True, DARK_GRAY), bg=GRAY_LIGHT, align=AL_R, border=BORDER_ALL)
    suma_m_formula = f"=SUM({_a(wt_method_start, C1+1)}:{_a(wt_method_end, C1+1)})"
    cell = ws.cell(row=r, column=C1 + 1, value=suma_m_formula)
    cell.font = _f(8, True, DARK_GRAY); cell.fill = fill(GRAY_LIGHT)
    cell.alignment = AL_C; cell.border = BORDER_ALL; cell.number_format = "0.0"
    # Red conditional formatting if method weights exceed 100%
    _red_fill2 = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    ws.conditional_formatting.add(
        f"{_gcl(C1+1)}{r}",
        CellIsRule(operator="greaterThan", formula=["100"], fill=_red_fill2)
    )
    merge(ws, r, C1 + 2, r, CE)
    r += 1

    # PRECIO OBJETIVO — big row
    ws.row_dimensions[r].height = 26
    wc(ws, r, C1, "  PRICE TARGET",
       font=_f(13, True, WHITE), bg=NAVY_DARK, align=AL_L, border=BORDER_ALL)
    merge(ws, r, C1, r, C1 + 1)

    contrib_sum = "+".join(contrib_cells)
    target_formula = f"={contrib_sum}"
    cell = ws.cell(row=r, column=C1 + 2, value=target_formula)
    cell.font = _f(13, True, WHITE); cell.fill = fill(NAVY_DARK)
    cell.alignment = AL_C; cell.border = BORDER_ALL; cell.number_format = '$#,##0.00'
    rows["target_price"] = (r, C1 + 2)
    merge(ws, r, C1 + 3, r, CE)
    wc(ws, r, C1 + 3, "", font=_f(9), bg=NAVY_DARK, align=AL_C, border=BORDER_ALL)
    r += 1

    # vs Precio Actual
    tp_r, tp_c = rows["target_price"]
    tp_ref = _a(tp_r, tp_c)

    ws.row_dimensions[r].height = 17
    wc(ws, r, C1, "vs Current Price", font=_f(9, True), bg=BLUE_TINT, align=AL_LI, border=BORDER_ALL)
    merge(ws, r, C1, r, C1 + 1)
    vs_formula = f"={tp_ref}/{price_ref}-1"
    cell = ws.cell(row=r, column=C1 + 2, value=vs_formula)
    cell.font = _f(9, True); cell.fill = fill(GREEN)
    cell.alignment = AL_C; cell.border = BORDER_ALL; cell.number_format = "0.0%"
    merge(ws, r, C1 + 3, r, CE)
    wc(ws, r, C1 + 3, "", font=_f(9), bg=GRAY_LIGHT, align=AL_C, border=BORDER_ALL)
    r += 1

    ws.row_dimensions[r].height = 17
    wc(ws, r, C1, "Implied Upside", font=_f(9, True), bg=BLUE_TINT, align=AL_LI, border=BORDER_ALL)
    merge(ws, r, C1, r, C1 + 1)
    upside_formula = f"={tp_ref}/{price_ref}-1"
    cell = ws.cell(row=r, column=C1 + 2, value=upside_formula)
    cell.font = _f(9, True); cell.fill = fill(BLUE_TINT)
    cell.alignment = AL_C; cell.border = BORDER_ALL; cell.number_format = "0.0%"
    merge(ws, r, C1 + 3, r, CE)
    wc(ws, r, C1 + 3, "", font=_f(9), bg=GRAY_LIGHT, align=AL_C, border=BORDER_ALL)
    r += 1

    spacer(ws, r); r += 1

    # Note
    ws.row_dimensions[r].height = 20
    wc(ws, r, C1,
       "  Yellow cells are editable — all other values are formula-driven. "
       "Adjust peer weights and method weights to see the impact on the price target.",
       font=_f(8, False, DARK_GRAY), bg=GRAY_LIGHT, align=AL_W)
    merge(ws, r, C1, r, CE)
    r += 1

    # Footer
    ws.row_dimensions[r].height = 14
    wc(ws, r, C1,
       "  Comparable company valuation. Does not constitute investment advice.",
       font=_f(8, False, MID_GRAY), bg=GRAY_LIGHT, align=AL_L)
    merge(ws, r, C1, r, CE)

    return ws
