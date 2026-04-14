"""
Excel Styles
Shared palette, fonts, fills, borders, alignments, and cell helpers.

Author: Joaquin Abondano w/ Claude Code
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Palette ───────────────────────────────────────────────────────────────────
NAVY_DARK  = "1F3864"
NAVY_MED   = "2C3E6B"
BLUE_ACC   = "4472C4"
GREEN      = "1A7741"
AMBER      = "9C6500"
RED        = "C0392B"
BLUE_TINT  = "D6E4F7"
GRAY_LIGHT = "F5F5F5"
GRAY_MED   = "D0D0D0"
WHITE      = "FFFFFF"
BLACK      = "000000"
DARK_GRAY  = "555555"
MID_GRAY   = "888888"

SIGNAL_BG   = {"bullish": GREEN,  "neutral": AMBER,     "bearish": RED}
SIGNAL_TEXT = {"bullish": "BULLISH  ▲", "neutral": "NEUTRAL  ●", "bearish": "BEARISH  ▼"}
ACTION_TEXT = {
    "buy":   "BUY  ↑",
    "hold":  "HOLD  →",
    "sell":  "SELL  ↓",
    "short": "SHORT  ↓",
    "cover": "COVER  ↑",
}


# ── Font helpers ──────────────────────────────────────────────────────────────
def _f(size=10, bold=False, color=BLACK, italic=False):
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)


# ── Fill helper ───────────────────────────────────────────────────────────────
def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


# ── Borders ───────────────────────────────────────────────────────────────────
_thin = Side(style="thin",   color=GRAY_MED)
_med  = Side(style="medium", color=NAVY_DARK)

BORDER_ALL = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BORDER_BOT = Border(bottom=_med)
BORDER_TOP = Border(top=_med)


# ── Alignments ────────────────────────────────────────────────────────────────
AL_L  = Alignment(horizontal="left",   vertical="center")
AL_LI = Alignment(horizontal="left",   vertical="center", indent=1)
AL_C  = Alignment(horizontal="center", vertical="center")
AL_R  = Alignment(horizontal="right",  vertical="center")
AL_W  = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
AL_WC = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ── Cell writer ───────────────────────────────────────────────────────────────
def wc(ws, row, col, value=None, *,
       font=None, bg=None, align=None, fmt=None, border=None):
    """Write and format a cell."""
    cell = ws.cell(row=row, column=col, value=value)
    if font   is not None: cell.font          = font
    if bg     is not None: cell.fill          = fill(bg)
    if align  is not None: cell.alignment     = align
    if fmt    is not None: cell.number_format = fmt
    if border is not None: cell.border        = border
    return cell


def merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def sec_hdr(ws, row, title, c1, c2, bg=NAVY_DARK):
    """Full-width navy section header spanning c1:c2."""
    cell = wc(ws, row, c1, f"  {title.upper()}",
              font=_f(10, True, WHITE), bg=bg, align=AL_L)
    merge(ws, row, c1, row, c2)
    ws.row_dimensions[row].height = 18
    return cell


def col_hdr(ws, row, labels, c_start, bg=NAVY_MED):
    """Row of column-header cells."""
    for i, lbl in enumerate(labels):
        wc(ws, row, c_start + i, lbl,
           font=_f(9, True, WHITE), bg=bg, align=AL_C, border=BORDER_ALL)
    ws.row_dimensions[row].height = 18


def spacer(ws, row, height=6):
    ws.row_dimensions[row].height = height


def hide_gridlines(ws):
    ws.sheet_view.showGridLines = False
